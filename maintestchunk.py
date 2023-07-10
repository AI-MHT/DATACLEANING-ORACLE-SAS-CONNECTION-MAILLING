import pandas as pd
import openpyxl

# Chunk processing function
def process_chunk(chunk):
    # Remove dashes from the NUM_SIN column
    chunk['NUM_SIN'] = chunk['NUM_SIN'].str.replace('-', '')

    # Replace values in the "DEGAT" column + name of column
    chunk['NATURE_SINISTRE'] = chunk['DEGAT'].replace({'MRC': 'MD', 'CRC': 'BI', 'MHRC': 'CASC', 'CHRC': 'CASC'})

    # Sum the columns 'RES_P_ACTUEL' and 'RES_H_ACTUEL' and create a new column 'RESERVE_GLOBALE'
    chunk['RESERVE_GLOBALE'] = chunk['RES_P_ACTUEL'] + chunk['RES_H_ACTUEL']

    # Sum the columns 'REG_P_CUMULE' and 'REG_H_CUMULE' and create a new column 'REGLEMENT_GLOBAL'
    chunk['REGLEMENT_GLOBAL'] = chunk['REG_P_CUMULE'] + chunk['REG_H_CUMULE']

    # Sum the columns 'REGLEMENT_GLOBAL' and 'RESERVE_GLOBALE' and create a new column 'CHARGE_GLOBALE'
    chunk['CHARGE_GLOBALE'] = chunk['REGLEMENT_GLOBAL'] + chunk['RESERVE_GLOBALE']

    # Replace empty or NaN values in 'CHARGE_GLOBALE', 'REGLEMENT_GLOBAL', and 'RESERVE_GLOBALE' with 0
    chunk['CHARGE_GLOBALE'] = chunk['CHARGE_GLOBALE'].fillna(0)
    chunk['REGLEMENT_GLOBAL'] = chunk['REGLEMENT_GLOBAL'].fillna(0)
    chunk['RESERVE_GLOBALE'] = chunk['RESERVE_GLOBALE'].fillna(0)

    # Convert 'DATE_CLOTURE' and 'DATE_OUVERTURE' columns to datetime type with the specified format
    chunk['DATE_CLOTURE'] = pd.to_datetime(chunk['DATE_CLOTURE'], format='%d/%m/%Y', errors='coerce').dt.date
    chunk['DATE_OUVERTURE'] = pd.to_datetime(chunk['DATE_OUVERTURE'], format='%d/%m/%Y', errors='coerce').dt.date

    # Add a new column 'STATUT_DOSS' based on the condition
    chunk['STATUT_DOSS'] = chunk.apply(lambda row: 'TERM' if pd.notna(row['DATE_CLOTURE']) and pd.notna(row['DATE_OUVERTURE']) and row['DATE_CLOTURE'] > row['DATE_OUVERTURE'] else 'OUV', axis=1)

    # Create a new column 'FLAG_INV' based on the condition
    chunk['FLAG_INV'] = chunk['LIBELLE_MOTIF'].astype(str).apply(lambda x: 1 if x.lower() == 'inventaire' else 0)

    # Convert 'DATE_VALIDATION' column to datetime type
    chunk['DATE_VALIDATION'] = pd.to_datetime(chunk['DATE_VALIDATION'], format='%d/%m/%Y', errors='coerce').dt.date

    # Add a new column 'FLAG_CID' and replace all cases with 'OKEY'
    chunk['FLAG_CID'] = 'O'

    # Add a new column 'SOURCE' and replace all cases with 'ABS'
    chunk['SOURCE'] = 'ABS'

    # Add a new column 'NATURE_RESPONSABILITE'' and replace all cases with 'N'
    chunk['NATURE_RESPONSABILITE'] = 'N'

    # Add a new column 'TYPE_RECOURS' and replace all cases with 'F'
    chunk['TYPE_RECOURS'] = 'F'

    # Add a new column 'TYPE_RECOURS' and replace all cases with 'F'
    chunk['REG_AV_1'] = 12000

    # Add a new column 'TYPE_RECOURS' and replace all cases with ''
    chunk['CHARGE_AV_1'] = -94399

    # Add a new column 'TYPE_RECOURS' and replace all cases with ''
    chunk['GAIN_REG'] = 98650

    # Add a new column 'TYPE_RECOURS' and replace all cases with ''
    chunk['GAIN_CHARGE'] = -98530

    # Add a new column 'TYPE_RECOURS' and replace all cases with ''
    chunk['SIN_CLOSED_AP_INV'] = 1

    # Add a new column 'TYPE_RECOURS' and replace all cases with 'F'
    chunk['LAG_FLAG_INV'] = 0

    chunk['STATUT'] = chunk['STATUT_DOSS']
    chunk['DEGAT'] = chunk['NATURE_SINISTRE']

    # Add a new column 'MOIS_VALID' to extract the month from 'DATE_VALIDATION'
    chunk['MOIS_VALID'] = chunk['DATE_VALIDATION'].apply(lambda x: x.month)

    # Add a new column 'ANNEE_VALID' to extract the year from 'DATE_VALIDATION'
    chunk['ANNEE_VALID'] = chunk['DATE_VALIDATION'].apply(lambda x: x.year)

    chunk['Y_VALID'] = chunk['ANNEE_VALID']

    # Create a new column 'KEY' by concatenating 'NUM_SIN' and 'NUMEACTE' with '||' in between
    chunk['KEY'] = chunk['NUM_SIN'] + '||' + chunk['NUMEACTE'].astype(str)

    # Group the data by 'KEY' and keep the first occurrence, dropping duplicates
    chunk = chunk.drop_duplicates(subset='KEY', keep='first')

    return chunk

# Read the CSV file in chunks
chunk_size = 1000000
filename = 'rrpaa.csv'
chunk_generator = pd.read_csv(filename, delimiter=';', chunksize=chunk_size)

# Process and save each chunk
output_filename = 'modified_file.xlsx'
with pd.ExcelWriter(output_filename) as writer:
    for i, chunk in enumerate(chunk_generator):
        processed_chunk = process_chunk(chunk)
        sheet_name = f'Sheet_{i + 1}'
        processed_chunk.to_excel(writer, sheet_name=sheet_name, index=False)

# Load the existing Excel file
existing_filename = 'apr.xlsx'
book = openpyxl.load_workbook(existing_filename)

# Create a new sheet with the modified data
sheet_name = 'DATA_INV_updated'
if sheet_name in book.sheetnames:
    book.remove(book[sheet_name])  # Remove the existing sheet
book.create_sheet(sheet_name, index=0)  # Create a new sheet
writer = pd.ExcelWriter(existing_filename, engine='openpyxl')
writer.book = book

# Write the modified data to the new sheet
for i, chunk in enumerate(chunk_generator):
    processed_chunk = process_chunk(chunk)
    sheet_name = f'Sheet_{i + 1}'
    processed_chunk.to_excel(writer, sheet_name=sheet_name, index=False)

# Save the changes and close the writer
writer.save()
writer.close()
