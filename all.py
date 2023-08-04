import pandas as pd
import openpyxl

# Read the CSV file and perform data modifications
data = pd.read_csv('rrpaa.csv', delimiter=';')
#data = pd.read_csv('rpa_csv.csv', delimiter=',', encoding='latin1')


# Remove dashes from the NUM_SIN column
data['NUM_SIN'] = data['NUM_SIN'].str.replace('-', '')

# Drop the "SOURCE" column from the DataFrame
# data = data.drop(columns=['IDENLIGN', 'ID_OPERATION', 'DATE_GENERATION', 'JOURNEE', 'PERIODE_DU', 'PERIODE_AU', 'ID_SINISTRE'])

# Replace values in the "DEGAT" column + name of column
data['NATURE_SINISTRE'] = data['DEGAT'].replace({'MRC': 'MD', 'CRC': 'BI', 'MHRC': 'CASC', 'CHRC': 'CASC'})

# Sum the columns 'RES_P_ACTUEL' and 'RES_H_ACTUEL' and create a new column 'RESERVE_GLOBALE'
data['RESERVE_GLOBALE'] = data['RES_P_ACTUEL'] + data['RES_H_ACTUEL']

# Sum the columns 'REG_P_CUMULE' and 'REG_H_CUMULE' and create a new column 'REGLEMENT_GLOBAL'
data['REGLEMENT_GLOBAL'] = data['REG_P_CUMULE'] + data['REG_H_CUMULE']

# Sum the columns 'REGLEMENT_GLOBAL' and 'RESERVE_GLOBALE' and create a new column 'CHARGE_GLOBALE'
data['CHARGE_GLOBALE'] = data['REGLEMENT_GLOBAL'] + data['RESERVE_GLOBALE']

# Replace empty or NaN values in 'CHARGE_GLOBALE', 'REGLEMENT_GLOBAL', and 'RESERVE_GLOBALE' with 0
data['CHARGE_GLOBALE'] = data['CHARGE_GLOBALE'].fillna(0)
data['REGLEMENT_GLOBAL'] = data['REGLEMENT_GLOBAL'].fillna(0)
data['RESERVE_GLOBALE'] = data['RESERVE_GLOBALE'].fillna(0)

# Drop the individual 'RES_P_ACTUEL', 'RES_H_ACTUEL', 'REG_P_CUMULE', and 'REG_H_CUMULE' columns if needed
data = data.drop(columns=['RES_P_ACTUEL', 'RES_H_ACTUEL', 'REG_P_CUMULE', 'REG_H_CUMULE'])

# Convert 'DATE_CLOTURE' and 'DATE_OUVERTURE' columns to datetime type with the specified format
data['DATE_CLOTURE'] = pd.to_datetime(data['DATE_CLOTURE'], format='%d/%m/%Y', errors='coerce').dt.date
data['DATE_OUVERTURE'] = pd.to_datetime(data['DATE_OUVERTURE'], format='%d/%m/%Y', errors='coerce').dt.date

# Add a new column 'STATUT_DOSS' based on the condition
data['STATUT_DOSS'] = data.apply(lambda row: 'TERM' if pd.notna(row['DATE_CLOTURE']) and pd.notna(row['DATE_OUVERTURE']) and row['DATE_CLOTURE'] > row['DATE_OUVERTURE'] else 'OUV', axis=1)

# Create a new column 'FLAG_INV' based on the condition
data['FLAG_INV'] = data['LIBELLE_MOTIF'].astype(str).apply(lambda x: 1 if x.lower() == 'inventaire' else 0)

# Convert 'DATE_VALIDATION' column to datetime type
data['DATE_VALIDATION'] = pd.to_datetime(data['DATE_VALIDATION']).dt.date

# Add a new column 'FLAG_CID' and replace all cases with 'OKEY'
data['FLAG_CID'] = 'O'

# Add a new column 'SOURCE' and replace all cases with 'ABS'
data['SOURCE'] = 'ABS'

# Add a new column 'NATURE_RESPONSABILITE'' and replace all cases with 'N'
data['NATURE_RESPONSABILITE'] = 'N'

# Add a new column 'TYPE_RECOURS' and replace all cases with 'F'
data['TYPE_RECOURS'] = 'F'

# Add a new column 'TYPE_RECOURS' and replace all cases with 'F'
data['REG_AV_1'] = 12000

# Add a new column 'TYPE_RECOURS' and replace all cases with ''
data['CHARGE_AV_1'] = -94399

# Add a new column 'TYPE_RECOURS' and replace all cases with ''
data['GAIN_REG'] = 98650

# Add a new column 'TYPE_RECOURS' and replace all cases with ''
data['GAIN_CHARGE'] = -98530

# Add a new column 'TYPE_RECOURS' and replace all cases with ''
data['SIN_CLOSED_AP_INV'] = 1

# Add a new column 'TYPE_RECOURS' and replace all cases with 'F'
data['LAG_FLAG_INV'] = 0

data['STATUT'] = data['STATUT_DOSS']
data['DEGAT'] = data['NATURE_SINISTRE']



# Add a new column 'MOIS_VALID' to extract the month from 'DATE_VALIDATION'
data['MOIS_VALID'] = data['DATE_VALIDATION'].apply(lambda x: x.month)

# Add a new column 'ANNEE_VALID' to extract the year from 'DATE_VALIDATION'
data['ANNEE_VALID'] = data['DATE_VALIDATION'].apply(lambda x: x.year)

data['Y_VALID'] = data['ANNEE_VALID']

# Create a new column 'KEY' by concatenating 'NUM_SIN' and 'NUMEACTE' with '||' in between
data['KEY'] = data['NUM_SIN'] + '||' + data['NUMEACTE'].astype(str)

# Group the data by 'KEY' and keep the first occurrence, dropping duplicates
data = data.drop_duplicates(subset='KEY', keep='first')


# Group the data by 'NUM_SIN' and 'NATURE_SINISTRE' and calculate the sum of 'RESERVE_GLOBALE'
grouped_data = data.groupby(['NUM_SIN', 'NATURE_SINISTRE'], as_index=False)['RESERVE_GLOBALE'].sum()

# Identify groups where 'NATURE_SINISTRE' is both 'BI' and 'MD' and the sum of 'RESERVE_GLOBALE' for either 'BI' or 'MD' is non-zero
mask = grouped_data.duplicated(subset=['NUM_SIN'], keep=False)
mixed_groups = grouped_data[mask].groupby('NUM_SIN').filter(lambda x: (x['NATURE_SINISTRE'].eq('BI').any() and x['RESERVE_GLOBALE'].ne(0).any()) and (x['NATURE_SINISTRE'].eq('MD').any() and x['RESERVE_GLOBALE'].ne(0).any()))

# Replace 'BI' and 'MD' with 'MIX' in the original DataFrame for the identified groups
data.loc[data['NUM_SIN'].isin(mixed_groups['NUM_SIN']) & (data['NATURE_SINISTRE'].isin(['BI', 'MD'])), 'NATURE_SINISTRE'] = 'MIX'

# Reorder the columns
data = data[['EXERCICE_SURVENANCE', 'NUM_SIN', 'NUMEACTE', 'DATE_VALIDATION', 'NATURE_SINISTRE','FLAG_CID','NATURE_RESPONSABILITE','TYPE_RECOURS', 'REGLEMENT_GLOBAL', 'RESERVE_GLOBALE', 'CHARGE_GLOBALE', 'DATE_OUVERTURE', 'DATE_CLOTURE', 'STATUT_DOSS', 'FLAG_INV','SOURCE', 'MOIS_VALID', 'ANNEE_VALID','REG_AV_1','CHARGE_AV_1','GAIN_REG','GAIN_CHARGE','SIN_CLOSED_AP_INV','LAG_FLAG_INV', 'KEY','STATUT','DEGAT','Y_VALID']]


# Save the modified data as an Excel file
data.to_excel('modified_file.xlsx', index=False)

# Load the existing Excel file
filename = 'apr.xlsx'
book = openpyxl.load_workbook(filename)

# Create a new sheet with the modified data
new_data = pd.read_excel('modified_file.xlsx')
sheet_name = 'DATA_INV_updated'
if sheet_name in book.sheetnames:
    book.remove(book[sheet_name])  # Remove the existing sheet
book.create_sheet(sheet_name, index=0)  # Create a new sheet
writer = pd.ExcelWriter(filename, engine='openpyxl')
writer.book = book
new_data.to_excel(writer, sheet_name=sheet_name, index=False)
writer.save()
writer.close()
#liste_modalite